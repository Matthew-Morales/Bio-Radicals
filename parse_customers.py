from openpyxl import load_workbook
import sqlite, sqlite3

database = "pythonsqlite.db"
conn = sqlite3.connect(database)

anemiaSheet = load_workbook("anemiasales.xlsx")
custSheet = load_workbook("rReport.xlsx")

anemia_sheet = anemiaSheet.active
cust_sheet = custSheet.active

headers = []
header_row = 0
count=0


def pullXOrders(sheet, ids, name, headerRow): # 8 for anemia, 1 for cust
    header_row = headerRow #starting header value
    headers = []   #list of headers to be returned
    mapping = {}   #customerID to sold-to party
    #remember to subtract 1 because counts start at 0 <3
    count = 0
    for each in sheet:
        #gets the column headers
        if (count == header_row):
            for e in each:
                headers.append(e.value)
            #print(headers)
        for i in range(len(headers)):
            if (str(each[ids].value)) not in mapping:
                if (str(each[ids].value)) != "#":
                    mapping[(str(each[ids].value))] = (str(each[name].value))
                else:
                    pass
        count += 1
    return mapping

def newOrderMerge(map1, map2):
    final = {}
    for key in map1:
        if key not in final:
            final[key] = map1[key]
    for key2 in map2:
        if key2 not in final:
            final[key2] = map2[key2]
    return final

def sameOrderMerge(map1, map2):
    for key in map2:
        if key not in map1:
            map1[key] = map2[key]

x = pullXOrders(cust_sheet, 0, 1, 1)
y = pullXOrders(cust_sheet, 2, 3, 1)


anemiaDict = pullXOrders(anemia_sheet, 4, 5, 7)


z = newOrderMerge(x, y)
customers = newOrderMerge(z, anemiaDict)
#print(len(z))
#print(len(customers))

for id in customers:
    print(id, customers[id])
    sqlite.create_customer(conn, (id, str(customers[id])))

conn.commit()


